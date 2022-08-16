import openpyxl

'''职业信息规范'''
'''针对未规范原因，匹配之前的未规范名单，如果之前存在，直接复用之前的原因'''

def matching1(path1, path_res):
    excel = openpyxl.load_workbook(path1)
    sheet = excel[excel.sheetnames[-1]]

    excel_res = openpyxl.load_workbook(path_res)
    sheet_res = excel_res[excel_res.sheetnames[-1]]

    list = []
    list_res = []

    for i in range(327):
        list.append(sheet['F' + str(i + 3)].value)

    for i in range(167):
        list_res.append(sheet_res['C' + str(i + 2)].value)

    for i in list:
        if i in list_res:
            index = list.index(i)
            index_res = list_res.index(i)
            print(sheet['K' + str(index + 3)].value)
            sheet_res['E' + str(index_res + 2)] = sheet['K' + str(index + 3)].value

    excel_res.save(path_res)

def matching2(path2, path_res):
    excel = openpyxl.load_workbook(path2)
    sheet = excel[excel.sheetnames[-1]]

    excel_res = openpyxl.load_workbook(path_res)
    sheet_res = excel_res[excel_res.sheetnames[-1]]

    list = []
    list_res = []

    for i in range(186):
        list.append(str(sheet['A' + str(i + 2)].value))
    print(list)
    print(len(list))

    for i in range(165):
        list_res.append(sheet_res['A' + str(i + 2)].value)
    print(list_res)
    print(len(list_res))

    for i in list:
        if i in list_res:
            index = list.index(i)
            index_res = list_res.index(i)
            print(sheet['B' + str(index + 2)].value)
            sheet_res['D' + str(index_res + 2)] = sheet['B' + str(index + 2)].value

    excel_res.save(path_res)

if __name__ == "__main__":

    path1 = "D:/projects/tools/excel_merge/excel1.xlsx"
    path2 = "D:/projects/tools/excel_merge/excel2.xlsx"
    path_res = "D:/projects/tools/excel_merge/result.xlsx"

    matching1(path1, path_res)
    # matching2(path2, path_res)
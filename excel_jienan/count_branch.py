import os
import openpyxl


def merge_excel(files, res_sheet1, res_sheet2, relate_dict):
    begin_count = 1
    counts = 2
    flag = 2

    for file in files:
        code = file.split('-')[0]
        file_name = file.split('-')[1]
        file_list = os.listdir(path + "/" + file)
        for f in file_list:
            listLOne = [file_name, code, {}]
            if "~$" in f:
                continue
            if "营运任务维护表" in f:
                print("正在统计： " + file_name + ": " + code)
                excel = openpyxl.load_workbook(path + "/" + file + "/" + f)
                sheet = excel[excel.sheetnames[-1]]
                count = len(sheet['A'])
                print(count)

                for i in range(2, count):
                    if "处理" in sheet['O2'].value:
                        mark = 'O'
                    elif "处理" in sheet['P2'].value:
                        mark = 'P'
                    elif "处理" in sheet['N2'].value:
                        mark = 'N'
                    else:
                        print("error!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
                        break

                    if sheet[mark + str(i)] == "待开立北京市场账户" or "未开立股转账户":
                        salepart = sheet['C'][i].value
                        branch_code = sheet['B'][i].value
                        branch_code = str(branch_code)
                        if len(branch_code) == 2:
                            branch_code = "00" + branch_code
                        if len(branch_code) == 3:
                            branch_code = "0" + branch_code

                        if branch_code not in relate_dict:
                            sale = salepart
                        else:
                            sale = relate_dict[branch_code]

                        if sale not in listLOne[2]:
                            listLOne[2][sale] = 1
                        else:
                            listLOne[2][sale] = listLOne[2][sale] + 1
                    else:
                        continue

        num = 0
        for key, value in listLOne[2].items():
            res_sheet1['A' + str(counts)] = file_name
            res_sheet1['B' + str(counts)] = code
            res_sheet1['C' + str(counts)] = key
            res_sheet1['D' + str(counts)] = value
            num += value
            counts += 1

        res_sheet2['A' + str(flag)] = file_name
        res_sheet2['B' + str(flag)] = code
        res_sheet2['C' + str(flag)] = num
        flag += 1

        print("---" + file_name + "共" + str(counts - begin_count - 1) + "个营业部")
        print("---" + file_name + "共" + str(num) + "个待开北京市场账户")
        begin_count = counts - 1
        print("------" + file_name + "finished!!\n")


if __name__ == "__main__":
    res_path = "D:/database/营业部统计/result_121-145.xlsx"
    res_excel = openpyxl.load_workbook(res_path)
    res_sheet1 = res_excel[res_excel.sheetnames[0]]
    res_sheet2 = res_excel[res_excel.sheetnames[-1]]

    relate_path = "D:/database/营业部统计/分支机构隶属关系.xlsx"
    relate_excel = openpyxl.load_workbook(relate_path)
    relate_sheet = relate_excel[relate_excel.sheetnames[0]]
    relate_dict = {}
    for i in range(1, len(relate_sheet['A'])):
        relate_dict[relate_sheet['A' + str(i)].value] = relate_sheet['B' + str(i)].value

    print(relate_dict)

    path = "D:/database/营业部统计/121-145"
    files = os.listdir(path)

    merge_excel(files, res_sheet1, res_sheet2, relate_dict)

    res_excel.save(res_path)
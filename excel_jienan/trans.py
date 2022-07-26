import os
import openpyxl


def merge_excel(files, res_sheet):
    begin_count = 1
    counts = 2

    for file in files:
        file_name = file.split('-')[1]
        begin_time = file.split('-')[-1]
        file_list = os.listdir(path + "/" + file)
        for f in file_list:
            if "~$" in f:
                continue
            if "营运任务维护表" in f:
                print("正在合并： " + file_name)
                excel = openpyxl.load_workbook(path + "/" + file + "/" + f)
                sheet = excel[excel.sheetnames[-1]]
                count = len(sheet['A'])

                for i in range(2, count):
                    content = sheet['L'][i].value

                    res_sheet['A'+str(counts)] = "3-股东代码"
                    res_sheet['B'+str(counts)] = content
                    res_sheet['C'+str(counts)] = "1-证券转板"
                    res_sheet['D'+str(counts)] = "1-完全禁止"
                    res_sheet['E'+str(counts)] = "1-默认"
                    res_sheet['F'+str(counts)] = "0-正常限制"
                    res_sheet['G'+str(counts)] = "1001-东贝B股股东证券转换禁止销户"
                    res_sheet['H'+str(counts)] = "1-禁止股东账户销户"
                    res_sheet['I'+str(counts)] = begin_time
                    res_sheet['J'+str(counts)] = "20891231"
                    res_sheet['K'+str(counts)] = "因" + file_name + "退市确权禁止销户"
                    counts += 1

        print("---" + file_name + "共" + str(counts - begin_count - 1) + "条数据")
        begin_count = counts - 1
        print("------" + file_name + "finished!!\n")


if __name__ == "__main__":
    res_path = "D:\database\截至0715确权.xlsx"
    res_excel = openpyxl.load_workbook(res_path)
    res_sheet = res_excel[res_excel.sheetnames[-1]]

    path = "D:\database\截至0715确权"
    files = os.listdir(path)

    merge_excel(files, res_sheet)

    res_excel.save(res_path)